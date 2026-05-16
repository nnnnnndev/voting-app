"""Loader for the per-AY faculty status spreadsheet.

The spreadsheet (faculty_status.xlsx in the project root) records each
faculty member's status for each academic year. Cells contain a status
keyword like "Admin", "GPD", "T/P Case", "Served", etc. Blank cells mean
the faculty member did not serve and had no special status that year.

Layout assumed:
  Row 1 (headers): ["", "Name", "AY 21-22", "AY 22-23", "AY 23-24", ...]
  Row 2+ : one row per faculty. Col 0 is a free-text marker for untenured
           faculty ("Pre-tenure" or similar), or blank for tenured. Col 1
           is the short/first name used in the sheet, which is mapped to
           the canonical faculty name via NAME_ALIASES below.

This loader is tolerant: missing file, missing rows, unknown statuses
all produce warnings rather than errors.
"""
from __future__ import annotations

import csv
import re
import warnings
from dataclasses import dataclass, field
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore[assignment]

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SHEET_PATH = PROJECT_ROOT / "faculty_status.xlsx"

# Filenames the loader will search for, in priority order.
SHEET_GLOB_PATTERNS = [
    "faculty_status.xlsx",
    "faculty_status.csv",
    "dpc_status*.xlsx",
    "dpc_status*.csv",
    "dpc history*.xlsx",
    "dpc history*.csv",
    "DPC history*.xlsx",
    "DPC history*.csv",
]

# When the file has no header row, the AY column labels come from
# config/rules.yaml (ay_columns). Exposed here for backward compatibility
# with code and tests that reference it directly.
def _ay_columns() -> list[str]:
    from . import app_config
    return list(app_config.load().ay_columns)


DEFAULT_HEADERLESS_AYS: list[str] = _ay_columns()


# Status catalog, hard-exclude set, alias map, and name aliases are loaded
# from config/rules.yaml at startup. The module-level constants below are
# populated by _refresh_from_config() and re-exported for backward
# compatibility with code that imported them.
STATUS_LABELS: dict[str, str] = {}
HARD_EXCLUDES: set[str] = set()
STATUS_ALIASES: dict[str, str] = {}
NAME_ALIASES: dict[str, str] = {}


def _refresh_from_config() -> None:
    """Pull status catalog, hard-exclude set, aliases, and name aliases
    from the YAML config. Called at import time and whenever the config
    cache is reset."""
    from . import app_config

    cfg = app_config.load()
    STATUS_LABELS.clear()
    STATUS_LABELS.update(cfg.status_labels)
    HARD_EXCLUDES.clear()
    HARD_EXCLUDES.update(cfg.hard_excludes)
    STATUS_ALIASES.clear()
    STATUS_ALIASES.update(cfg.status_aliases)
    NAME_ALIASES.clear()
    NAME_ALIASES.update(cfg.name_aliases)


_refresh_from_config()


@dataclass(frozen=True)
class StatusEntry:
    ay: str        # e.g. "AY 23-24"
    code: str      # canonical code, e.g. "GPD"
    raw: str       # original cell text, for display fallback

    @property
    def label(self) -> str:
        return STATUS_LABELS.get(self.code, self.raw)

    @property
    def display(self) -> str:
        return f"{self.label} {self.ay}"


@dataclass
class FacultyStatusSheet:
    ays: list[str] = field(default_factory=list)
    # by canonical name -> list of StatusEntry (one per non-blank AY)
    by_name: dict[str, list[StatusEntry]] = field(default_factory=dict)
    # by canonical name -> True if marked untenured in left column
    untenured: dict[str, bool] = field(default_factory=dict)
    # names found in sheet but not in CSV (warning surface)
    unmatched: list[str] = field(default_factory=list)

    @property
    def active_ay(self) -> str:
        """Most recent AY column in the sheet. The vote app uses this as
        the year being composed unless overridden."""
        return self.ays[-1] if self.ays else ""

    def statuses_for(self, canonical_name: str, ay: str) -> list[StatusEntry]:
        return [e for e in self.by_name.get(canonical_name, []) if e.ay == ay]

    def is_hard_excluded(self, canonical_name: str, ay: str) -> tuple[bool, str]:
        for e in self.statuses_for(canonical_name, ay):
            if e.code in HARD_EXCLUDES:
                return True, e.label
        return False, ""

    def is_untenured(self, canonical_name: str) -> bool:
        return self.untenured.get(canonical_name, False)

    def recent_statuses(self, canonical_name: str, n: int = 4) -> list[StatusEntry]:
        entries = self.by_name.get(canonical_name, [])
        # Sort by AY descending (most recent first).
        entries = sorted(entries, key=lambda e: e.ay, reverse=True)
        return entries[:n]


def _normalize(text: str) -> str:
    s = re.sub(r"\(.*?\)", "", text or "")  # drop parenthetical notes like "(?)"
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def _parse_ay(header: str) -> str | None:
    if not header:
        return None
    h = str(header).strip()
    # Accept "AY 21-22", "21-22", "2021-22", "AY 2021-2022", etc.
    m = re.search(r"(\d{2,4})\s*[-–]\s*(\d{2,4})", h)
    if not m:
        return None
    a, b = m.group(1), m.group(2)
    if len(a) == 4:
        a = a[-2:]
    if len(b) == 4:
        b = b[-2:]
    return f"AY {a}-{b}"


def find_sheet_path(root: Path | None = None) -> Path | None:
    """Locate the status sheet by trying known filename patterns."""
    root = root or PROJECT_ROOT
    for pat in SHEET_GLOB_PATTERNS:
        matches = sorted(root.glob(pat), key=lambda p: p.stat().st_mtime, reverse=True)
        if matches:
            return matches[0]
    return None


def _read_rows(path: Path) -> list[list]:
    """Return list-of-list of cell values for CSV or XLSX, in row order."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with open(path, encoding="utf-8-sig", newline="") as f:
            return [list(r) for r in csv.reader(f)]
    if suffix in (".xlsx", ".xlsm"):
        if load_workbook is None:
            return []
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        ws = wb.active
        return [list(r) for r in ws.iter_rows(values_only=True)]
    return []


def _detect_header(rows: list[list]) -> bool:
    """True if row 0 looks like a header (contains 'Name' or AY labels)."""
    if not rows:
        return False
    row0 = [str(c or "").strip() for c in rows[0]]
    for c in row0:
        if c.lower() == "name":
            return True
        if _parse_ay(c):
            return True
    return False


def load_sheet(
    sheet_path: Path | None = None,
    known_names: set[str] | None = None,
) -> FacultyStatusSheet:
    """Load the status sheet (CSV or XLSX). Returns an empty sheet if the
    file is missing; the app remains functional without it."""
    if sheet_path is None:
        sheet_path = find_sheet_path() or SHEET_PATH
    sheet = FacultyStatusSheet()
    if not sheet_path or not sheet_path.exists():
        return sheet

    rows = _read_rows(sheet_path)
    if not rows:
        return sheet

    has_header = _detect_header(rows)

    name_col: int | None
    untenured_col: int | None
    ay_cols: list[tuple[int, str]]

    if has_header:
        header = rows[0]
        name_col = None
        ay_cols = []
        for i, h in enumerate(header):
            h_str = str(h or "").strip()
            if h_str.lower() == "name":
                name_col = i
                continue
            ay = _parse_ay(h_str)
            if ay:
                ay_cols.append((i, ay))
        if name_col is None:
            for i, h in enumerate(header):
                if h:
                    name_col = i
                    break
        untenured_col = 0 if name_col != 0 else None
        data_rows = rows[1:]
        if name_col is None or not ay_cols:
            warnings.warn(
                f"{sheet_path.name}: could not find Name and AY columns in header"
            )
            return sheet
    else:
        # Headerless format: col 0 = untenured marker, col 1 = name,
        # cols 2..N = AY status. Labels come from config/rules.yaml.
        name_col = 1
        untenured_col = 0
        ay_labels = _ay_columns()
        n_ay = len(ay_labels)
        ay_cols = [(2 + i, ay_labels[i]) for i in range(n_ay)]
        data_rows = rows

    sheet.ays = [ay for _, ay in ay_cols]

    for row in data_rows:
        if not row:
            continue
        col0 = str(row[0] or "").strip() if len(row) > 0 else ""
        # Stop at legend section.
        if col0.upper().startswith("LEGEND"):
            break
        raw_name = row[name_col] if name_col < len(row) else None
        if not raw_name:
            continue
        raw_name = str(raw_name).strip()
        # Skip legend description rows (Status keyword in col 0, sentence in col 1).
        if col0 in STATUS_LABELS.values() or _normalize(col0) in STATUS_ALIASES:
            continue
        canonical = NAME_ALIASES.get(raw_name, raw_name)
        if known_names is not None and canonical not in known_names:
            matches = [n for n in known_names if n.split()[0] == raw_name]
            if len(matches) == 1:
                canonical = matches[0]
            else:
                sheet.unmatched.append(raw_name)
                continue

        if untenured_col is not None and untenured_col < len(row):
            marker = row[untenured_col]
            if marker and str(marker).strip():
                # Only record when explicitly marked. Blank cells should
                # leave rank-based classification untouched.
                sheet.untenured[canonical] = True

        entries = []
        for col_idx, ay in ay_cols:
            if col_idx >= len(row):
                continue
            cell = row[col_idx]
            if cell is None:
                continue
            raw = str(cell).strip()
            if not raw:
                continue
            norm = _normalize(raw)
            code = STATUS_ALIASES.get(norm)
            if code is None:
                norm2 = re.sub(r"[^a-z0-9 ]", "", norm)
                code = STATUS_ALIASES.get(norm2)
            if code is None:
                warnings.warn(
                    f"{sheet_path.name}: unknown status '{raw}' "
                    f"for {raw_name} in {ay}"
                )
                continue
            entries.append(StatusEntry(ay=ay, code=code, raw=raw))
        sheet.by_name[canonical] = entries

    return sheet


_cache: FacultyStatusSheet | None = None


def get_sheet(known_names: set[str] | None = None) -> FacultyStatusSheet:
    global _cache
    if _cache is None:
        _cache = load_sheet(known_names=known_names)
    return _cache


def reset_cache() -> None:
    global _cache
    _cache = None
